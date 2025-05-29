import streamlit as st
import pandas as pd
import re
from io import BytesIO
import unicodedata
from io import StringIO
from difflib import SequenceMatcher
from collections import defaultdict
import openpyxl
from openpyxl.styles import PatternFill
import random

# Strict email regex to avoid false positives
email_pattern = re.compile(
    r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
)

def is_email(value):
    value = str(value).strip()
    return bool(email_pattern.match(value))

def detect_string_columns(df):
    string_cols = []
    for col in df.columns:
        series = df[col].dropna()
        # Filter strings
        string_values = series[series.apply(lambda x: isinstance(x, str))]
        if not string_values.empty:
            # Exclude if any looks like email
            if not string_values.map(is_email).any():
                string_cols.append(col)
        # Check if column has any string with alphabetic char
        has_text = series.astype(str).apply(lambda x: any(c.isalpha() for c in x)).any()
        # Exclude columns that contain emails
        contains_email = series.astype(str).map(is_email).any()
        # Exclude numeric-only columns
        is_numeric = pd.api.types.is_numeric_dtype(df[col])
        if has_text and not contains_email and not is_numeric:
            string_cols.append(col)
    return string_cols

def detect_numeric_columns(df):
    """Detect columns that likely contain numeric data (quantities, prices, etc.)"""
    numeric_cols = []
    for col in df.columns:
        if pd.api.types.is_numeric_dtype(df[col]):
            numeric_cols.append(col)
        else:
            # Check if column contains numeric-like strings
            sample_values = df[col].dropna().astype(str).head(100)
            numeric_count = 0
            for val in sample_values:
                # Remove common non-numeric characters and check if remainder is numeric
                cleaned = re.sub(r'[,$\s]', '', str(val))
                try:
                    float(cleaned)
                    numeric_count += 1
                except:
                    pass
            
            if numeric_count / len(sample_values) > 0.7:  # 70% numeric-like values
                numeric_cols.append(col)
    
    return numeric_cols

def detect_categorical_columns(df, exclude_clusters=True):
    """Detect columns suitable for grouping/categorization"""
    categorical_cols = []
    for col in df.columns:
        if exclude_clusters and '_cluster' in col:
            continue
        
        # Skip numeric columns
        if pd.api.types.is_numeric_dtype(df[col]):
            continue
            
        # Check unique value ratio
        unique_ratio = df[col].nunique() / len(df)
        
        # Good categorical columns have reasonable number of unique values
        if 0.01 <= unique_ratio <= 0.3:  # Between 1% and 30% unique values
            categorical_cols.append(col)
    
    return categorical_cols

def safe_numeric_conversion(series):
    """Safely convert a series to numeric, handling common formats"""
    def convert_value(val):
        if pd.isna(val):
            return 0
        
        val_str = str(val).strip()
        
        # Remove common non-numeric characters
        cleaned = re.sub(r'[,$\s]', '', val_str)
        
        try:
            return float(cleaned)
        except:
            return 0
    
    return series.apply(convert_value)

def convert_df_to_csv_bytes(df):
    return df.to_csv(index=False).encode('utf-8')

def clean_pin(value):
    if pd.isna(value):
        return value
    value = str(value)
    # Remove "pin-" prefix, case-insensitive
    value = re.sub(r'pin-', '', value, flags=re.IGNORECASE).strip()
    # Extract first group of 6 digits
    match = re.search(r'\b(\d{6})\b', value)
    return match.group(1) if match else value

def standardize_value(val, col_name=""):
    if pd.isna(val):
        return val
    
    val_str = str(val)

    if val_str.strip() == "":
        return val_str
    
    if "pin" in col_name.lower():
        return clean_pin(val)

    val_str = unicodedata.normalize('NFKD', val_str).encode('ascii', 'ignore').decode('utf-8')
    val_str = val_str.lower()
    val_str = val_str.strip()
    val_str = re.sub(r'\s+', ' ', val_str)

    return val_str

def standardize_dataframe(df, string_cols):
    df = df.copy()
    for col in string_cols:
        df[col] = df[col].apply(lambda x: standardize_value(x, col_name=col))
    return df

# NEW CLUSTERING FUNCTIONS
def extract_core_product_name(text):
    """Extract the core product name by preserving important product codes"""
    if pd.isna(text):
        return ""
    
    original_text = str(text).strip()
    text = original_text.lower()
    
    # First, extract important product codes before removing parentheses
    # Look for patterns like (ar-740), (ar-825h), (pq0015066), etc.
    product_codes = []
    
    # Extract alphanumeric codes with hyphens (like ar-740, ar-825h)
    code_matches = re.findall(r'\(([a-z]{2,3}-?\d+[a-z]*)\)', text)
    product_codes.extend(code_matches)
    
    # Extract other product codes (like pq0015066)
    other_codes = re.findall(r'\(([a-z]{2}\d+)\)', text)
    product_codes.extend(other_codes)
    
    # Remove descriptions in parentheses but keep the main text structure
    text = re.sub(r'\([^)]*\)', '', text)
    
    # Clean up extra spaces
    text = re.sub(r'\s+', ' ', text).strip()
    
    # Extract the base product name (like "acm", "lipolan f", etc.)
    base_name = ""
    
    # Try to match common patterns
    if re.match(r'^[a-z]+\s*[a-z]*', text):
        # Extract first 1-2 words as base name
        words = text.split()
        if len(words) >= 2:
            base_name = f"{words[0]} {words[1]}"
        else:
            base_name = words[0] if words else ""
    else:
        base_name = text
    
    # Combine base name with the most specific product code
    if product_codes:
        # Prioritize codes with hyphens and letters (more specific)
        specific_codes = [code for code in product_codes if '-' in code and any(c.isalpha() for c in code)]
        if specific_codes:
            return f"{base_name} {specific_codes[0]}"
        else:
            return f"{base_name} {product_codes[0]}"
    
    return base_name

def similarity_score(str1, str2):
    """Calculate similarity between two strings"""
    return SequenceMatcher(None, str1, str2).ratio()

def cluster_product_names(series, similarity_threshold=0.8):
    """Cluster similar product names together with better product code handling"""
    if series.empty:
        return pd.Series([], dtype=str)
    
    # Get unique values and their core names
    unique_values = series.dropna().unique()
    core_names = {val: extract_core_product_name(val) for val in unique_values}
    
    # Create direct mapping - each unique core name becomes a cluster
    clusters = {}
    
    for val, core in core_names.items():
        if core and core.strip():  # Only process non-empty core names
            clusters[val] = core
        else:
            # Fallback for items without clear core names
            clusters[val] = str(val).lower().strip()
    
    # Map the series values to cluster names
    return series.map(lambda x: clusters.get(x, str(x).lower().strip() if pd.notna(x) else x))

def add_cluster_column(df, column_name):
    """Add a cluster column for the specified column"""
    if column_name not in df.columns:
        return df
    
    df_copy = df.copy()
    cluster_col_name = f"{column_name}_cluster"
    df_copy[cluster_col_name] = cluster_product_names(df_copy[column_name])
    
    return df_copy

def generate_colors(n):
    """Generate n distinct colors for clusters"""
    colors = [
        'FFE6E6', 'E6F3FF', 'E6FFE6', 'FFF0E6', 'F0E6FF',
        'FFFFE6', 'FFE6F0', 'E6FFFF', 'F0FFE6', 'FFE6CC',
        'E6E6FF', 'CCFFE6', 'FFE6B3', 'E6CCFF', 'B3FFE6',
        'FFB3E6', 'B3E6FF', 'E6FFB3', 'FFB3CC', 'CCFFB3',
        'FFD700', 'FFB6C1', '98FB98', 'DDA0DD', 'F0E68C',
        'FFA07A', '20B2AA', 'FFE4B5', 'D3D3D3', 'F5DEB3'
    ]
    
    if n <= len(colors):
        return colors[:n]
    else:
        # Generate additional random colors if needed
        additional_colors = []
        for _ in range(n - len(colors)):
            color = f"{random.randint(200, 255):02X}{random.randint(200, 255):02X}{random.randint(200, 255):02X}"
            additional_colors.append(color)
        return colors + additional_colors

def create_colored_excel(df, cluster_column):
    """Create an Excel file with color-coded clusters"""
    cluster_col = f"{cluster_column}_cluster"
    
    if cluster_col not in df.columns:
        return None
    
    # Sort by cluster to group similar items together
    df_sorted = df.sort_values(by=cluster_col).reset_index(drop=True)
    
    # Get unique clusters and assign colors
    unique_clusters = df_sorted[cluster_col].unique()
    colors = generate_colors(len(unique_clusters))
    cluster_colors = dict(zip(unique_clusters, colors))
    
    # Create Excel file in memory
    output = BytesIO()
    
    # Write to Excel
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Write main data sheet
        df_sorted.to_excel(writer, sheet_name='Clustered_Data', index=False)
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Clustered_Data']
        
        # Apply colors to rows based on clusters
        cluster_col_idx = df_sorted.columns.get_loc(cluster_col) + 1  # +1 for Excel 1-based indexing
        
        for row in range(2, len(df_sorted) + 2):  # Start from row 2 (after header)
            cluster_value = df_sorted.iloc[row-2][cluster_col]
            color_hex = cluster_colors.get(cluster_value, 'FFFFFF')
            fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')
            
            # Apply color to entire row
            for col in range(1, len(df_sorted.columns) + 1):
                worksheet.cell(row=row, column=col).fill = fill
        
        # Create a summary sheet with cluster information
        cluster_summary = df_sorted.groupby(cluster_col).size().reset_index(name='Count')
        cluster_summary['Color'] = cluster_summary[cluster_col].map(cluster_colors)
        cluster_summary.to_excel(writer, sheet_name='Cluster_Summary', index=False)
        
        # Apply colors to summary sheet
        summary_sheet = writer.sheets['Cluster_Summary']
        for row in range(2, len(cluster_summary) + 2):
            cluster_value = cluster_summary.iloc[row-2][cluster_col]
            color_hex = cluster_colors.get(cluster_value, 'FFFFFF')
            fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')
            
            for col in range(1, len(cluster_summary.columns) + 1):
                summary_sheet.cell(row=row, column=col).fill = fill
    
    output.seek(0)
    return output.getvalue()

def perform_cluster_analysis(df, cluster_col, analysis_type, target_col=None, group_by_col=None, selected_clusters=None):
    """Perform various types of analysis on clustered data"""
    
    if cluster_col not in df.columns:
        return None, "Cluster column not found"
    
    # Filter by selected clusters if specified
    if selected_clusters:
        df_filtered = df[df[cluster_col].isin(selected_clusters)]
    else:
        df_filtered = df
    
    if df_filtered.empty:
        return None, "No data found for selected clusters"
    
    try:
        if analysis_type == "cluster_summary":
            # Basic cluster summary
            result = df_filtered.groupby(cluster_col).agg({
                cluster_col: 'count'
            }).rename(columns={cluster_col: 'Total_Records'})
            
            if target_col and target_col in df_filtered.columns:
                numeric_data = safe_numeric_conversion(df_filtered[target_col])
                df_temp = df_filtered.copy()
                df_temp[f'{target_col}_numeric'] = numeric_data
                
                summary = df_temp.groupby(cluster_col)[f'{target_col}_numeric'].agg([
                    'sum', 'mean', 'count'
                ]).round(2)
                summary.columns = [f'{target_col}_Total', f'{target_col}_Average', f'{target_col}_Count']
                
                result = pd.concat([result, summary], axis=1)
            
            return result, "Analysis completed successfully"
        
        elif analysis_type == "top_clusters":
            if not target_col or target_col not in df_filtered.columns:
                return None, "Target column required for top clusters analysis"
            
            numeric_data = safe_numeric_conversion(df_filtered[target_col])
            df_temp = df_filtered.copy()
            df_temp[f'{target_col}_numeric'] = numeric_data
            
            result = df_temp.groupby(cluster_col)[f'{target_col}_numeric'].sum().sort_values(ascending=False).head(10)
            result = result.to_frame(f'Total_{target_col}')
            
            return result, "Top clusters analysis completed"
        
        elif analysis_type == "cluster_by_category":
            if not group_by_col or group_by_col not in df_filtered.columns:
                return None, "Group by column required for categorical analysis"
            
            if target_col and target_col in df_filtered.columns:
                numeric_data = safe_numeric_conversion(df_filtered[target_col])
                df_temp = df_filtered.copy()
                df_temp[f'{target_col}_numeric'] = numeric_data
                
                result = df_temp.groupby([cluster_col, group_by_col])[f'{target_col}_numeric'].sum().unstack(fill_value=0)
            else:
                result = df_filtered.groupby([cluster_col, group_by_col]).size().unstack(fill_value=0)
            
            return result, "Categorical analysis completed"
        
        elif analysis_type == "detailed_breakdown":
            if not group_by_col or group_by_col not in df_filtered.columns:
                return None, "Group by column required for detailed breakdown"
            
            result_list = []
            
            for cluster in df_filtered[cluster_col].unique():
                cluster_data = df_filtered[df_filtered[cluster_col] == cluster]
                
                breakdown = cluster_data.groupby(group_by_col).agg({
                    cluster_col: 'count'
                }).rename(columns={cluster_col: 'Record_Count'})
                
                if target_col and target_col in df_filtered.columns:
                    numeric_data = safe_numeric_conversion(cluster_data[target_col])
                    cluster_data_temp = cluster_data.copy()
                    cluster_data_temp[f'{target_col}_numeric'] = numeric_data
                    
                    summary = cluster_data_temp.groupby(group_by_col)[f'{target_col}_numeric'].sum()
                    breakdown[f'Total_{target_col}'] = summary
                
                breakdown['Cluster'] = cluster
                result_list.append(breakdown.reset_index())
            
            if result_list:
                result = pd.concat(result_list, ignore_index=True)
                return result, "Detailed breakdown completed"
            else:
                return None, "No data to analyze"
        
    except Exception as e:
        return None, f"Analysis error: {str(e)}"
    
    return None, "Unknown analysis type"

# Streamlit UI starts here
st.title("Automatic String Column Standardizer with Clustering")

uploaded_file = st.file_uploader("Upload your CSV file", type=["csv"])

if uploaded_file:
    # Read CSV with fallback encoding
    try:
        df = pd.read_csv(uploaded_file)
    except UnicodeDecodeError:
        uploaded_file.seek(0)
        df = pd.read_csv(uploaded_file, encoding='ISO-8859-1')

    st.subheader("Original Data Sample")
    st.dataframe(df.head(10))

    string_cols = detect_string_columns(df)
    st.write(f"**Detected string columns (to standardize):** {string_cols}")

    if st.button("Standardize String Columns"):
        df_clean = standardize_dataframe(df, string_cols)
        st.subheader("Standardized Data Sample")
        st.dataframe(df_clean.head(10))

        # Prepare CSV for download
        csv = df_clean.to_csv(index=False).encode('utf-8')
        st.download_button(
            label="Download Standardized CSV",
            data=csv,
            file_name="standardized_output.csv",
            mime="text/csv"
        )
        
        # Store standardized data in session state for clustering
        st.session_state['df_standardized'] = df_clean
        st.session_state['string_cols'] = string_cols

# Clustering section (only show if standardized data exists)
if 'df_standardized' in st.session_state:
    st.subheader("Product Name Clustering")
    st.write("Select a column to cluster similar product names:")
    
    df_std = st.session_state['df_standardized']
    cluster_column = st.selectbox(
        "Choose column for clustering:",
        options=st.session_state['string_cols'],
        key="cluster_column_select"
    )
    
    if st.button("Create Clusters"):
        df_clustered = add_cluster_column(df_std, cluster_column)
        
        # Store clustered data in session state
        st.session_state['df_clustered'] = df_clustered
        st.session_state['cluster_column_name'] = cluster_column
        
        st.subheader("Data with Clusters")
        # Show original, standardized, and clustered columns side by side
        display_cols = [cluster_column, f"{cluster_column}_cluster"]
        if cluster_column in df_clustered.columns:
            st.dataframe(df_clustered[display_cols].head(20))
        
        # Show cluster summary
        cluster_col = f"{cluster_column}_cluster"
        if cluster_col in df_clustered.columns:
            cluster_counts = df_clustered[cluster_col].value_counts()
            st.subheader("Cluster Summary")
            st.write(f"Total unique clusters: {len(cluster_counts)}")
            st.dataframe(cluster_counts.head(10).to_frame("Count"))
        
        # Download clustered data as CSV
        csv_clustered = df_clustered.to_csv(index=False).encode('utf-8')
        st.download_button(
            label="Download Data with Clusters (CSV)",
            data=csv_clustered,
            file_name="clustered_output.csv",
            mime="text/csv"
        )

# Show clustered data and Excel export if clustering has been done
if 'df_clustered' in st.session_state:
    df_clustered = st.session_state['df_clustered']
    cluster_column = st.session_state['cluster_column_name']
    
    # Show the clustered data again
    st.subheader("Clustered Data (Persistent)")
    display_cols = [cluster_column, f"{cluster_column}_cluster"]
    st.dataframe(df_clustered[display_cols].head(20))
    
    # Show cluster summary
    cluster_col = f"{cluster_column}_cluster"
    cluster_counts = df_clustered[cluster_col].value_counts()
    st.write(f"**Total unique clusters:** {len(cluster_counts)}")
    
    # Color-Coded Excel Export Section
    st.subheader("Color-Coded Excel Export")
    st.write("Generate an Excel file where each cluster is color-coded and grouped together:")
    
    if st.button("Generate Color-Coded Excel", key="excel_export"):
        with st.spinner("Creating color-coded Excel file..."):
            excel_data = create_colored_excel(df_clustered, cluster_column)
            
            if excel_data:
                st.session_state['excel_data'] = excel_data
                st.session_state['excel_ready'] = True
                st.success("✅ Excel file generated successfully!")
    
    # Show download button if Excel is ready
    if st.session_state.get('excel_ready', False) and 'excel_data' in st.session_state:
        st.download_button(
            label="📊 Download Color-Coded Excel File",
            data=st.session_state['excel_data'],
            file_name="clustered_data_colored.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        # Show preview of what the Excel will contain
        cluster_col = f"{cluster_column}_cluster"
        preview_data = df_clustered.groupby(cluster_col).size().reset_index(name='Row_Count')
        st.write("**Excel File Contents:**")
        st.write("- **Clustered_Data** sheet: All rows grouped by cluster with color coding")
        st.write("- **Cluster_Summary** sheet: Summary of clusters with counts")
        st.write("**Cluster Distribution:**")
        st.dataframe(preview_data.head(10))

    # DATA ANALYTICS SECTION
    st.subheader("📊 Data Analytics & Insights")
    st.write("Query your clustered data to get analytical insights:")
    
    # Detect column types for better user experience
    numeric_cols = detect_numeric_columns(df_clustered)
    categorical_cols = detect_categorical_columns(df_clustered)
    
    # Analytics interface
    col1, col2 = st.columns(2)
    
    with col1:
        st.write("**Available Numeric Columns (for calculations):**")
        st.write(numeric_cols if numeric_cols else "No numeric columns detected")
        
    with col2:
        st.write("**Available Categorical Columns (for grouping):**")
        st.write(categorical_cols if categorical_cols else "No categorical columns detected")
    
    # Analysis type selection
    analysis_type = st.selectbox(
        "Select Analysis Type:",
        [
            "cluster_summary",
            "top_clusters", 
            "cluster_by_category",
            "detailed_breakdown"
        ],
        format_func=lambda x: {
            "cluster_summary": "📈 Cluster Summary (Total records, sums, averages)",
            "top_clusters": "🏆 Top Clusters (Ranked by selected metric)",
            "cluster_by_category": "📊 Cross-Analysis (Clusters vs Categories)",
            "detailed_breakdown": "🔍 Detailed Breakdown (Complete analysis by category)"
        }[x]
    )
    
    # Dynamic input fields based on analysis type
    target_col = None
    group_by_col = None
    selected_clusters = None
    
    if analysis_type in ["cluster_summary", "top_clusters", "cluster_by_category", "detailed_breakdown"]:
        if numeric_cols:
            target_col = st.selectbox(
                "Select Numeric Column for Calculations (optional):",
                ["None"] + numeric_cols
            )
            target_col = None if target_col == "None" else target_col
    
    if analysis_type in ["cluster_by_category", "detailed_breakdown"]:
        if categorical_cols:
            group_by_col = st.selectbox(
                "Group By Column:",
                categorical_cols
            )
    
    # Cluster selection
    all_clusters = sorted(df_clustered[cluster_col].unique())
    selected_clusters = st.multiselect(
        "Select Specific Clusters (leave empty for all):",
        all_clusters,
        default=[]
    )
    
    if not selected_clusters:
        selected_clusters = None
    
    # Run analysis button
    if st.button("🔍 Run Analysis", key="run_analysis"):
        with st.spinner("Analyzing data..."):
            result, message = perform_cluster_analysis(
                df_clustered, 
                cluster_col, 
                analysis_type, 
                target_col, 
                group_by_col, 
                selected_clusters
            )
            
            if result is not None:
                st.success(message)
                st.subheader("Analysis Results")
                st.dataframe(result)
                
                # Download results
                csv_results = result.to_csv().encode('utf-8')
                st.download_button(
                    label="📥 Download Analysis Results",
                    data=csv_results,
                    file_name=f"analysis_{analysis_type}.csv",
                    mime="text/csv"
                )
                
                # Store results in session state
                st.session_state['analysis_results'] = result
                st.session_state['analysis_type'] = analysis_type
                
            else:
                st.error(f"Analysis failed: {message}")
    
    # Quick insights section
    if 'analysis_results' in st.session_state:
        st.subheader("💡 Quick Insights")
        result = st.session_state['analysis_results']
        analysis_type = st.session_state['analysis_type']
        
        if analysis_type == "cluster_summary":
            st.write(f"**Total Clusters Analyzed:** {len(result)}")
            if 'Total_Records' in result.columns:
                st.write(f"**Largest Cluster:** {result['Total_Records'].idxmax()} ({result['Total_Records'].max()} records)")
            
            if target_col and f'{target_col}_Total' in result.columns:
                st.write(f"**Highest {target_col} Total:** {result[f'{target_col}_Total'].idxmax()} ({result[f'{target_col}_Total'].max():,.2f})")
        
        elif analysis_type == "top_clusters":
            st.write(f"**Top Performing Cluster:** {result.index[0]} ({result.iloc[0, 0]:,.2f})")
            st.write(f"**Bottom Performing Cluster:** {result.index[-1]} ({result.iloc[-1, 0]:,.2f})")


